from flask import Flask, render_template, request, redirect, url_for
import os
from openpyxl import Workbook, load_workbook
import resend

app = Flask(__name__)

EXCEL_FILE = "registrations.xlsx"

# Create Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "Student Type", "Ticket Type", "Price",
        "Name", "Email", "Phone", "Institution",
        "Panel Preference", "Dietary Preference", "Special Requirements"
    ])
    wb.save(EXCEL_FILE)

# Resend setup
resend.api_key = os.getenv("RESEND_API_KEY")
EMAIL_SENDER = os.getenv("EMAIL_SENDER")
EMAIL_RECEIVER = os.getenv("EMAIL_RECEIVER")

def send_email(to_email, subject, body_html):
    try:
        resend.Emails.send({
            "from": EMAIL_SENDER,
            "to": [to_email],
            "subject": subject,
            "html": body_html
        })
        return True
    except Exception as e:
        print(f"Email sending failed: {e}")
        return False

@app.route("/", methods=["GET", "POST"])
@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        student_type = request.form["student_type"]
        ticket_type = request.form["ticket_type"]
        price = request.form["price"]
        name = request.form["name"]
        email = request.form["email"]
        phone = request.form["phone"]
        institution = request.form["institution"]
        panel = request.form["panel"]
        diet = request.form.get("diet", "")
        special = request.form.get("special", "")

        # Save to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append([
            student_type, ticket_type, price,
            name, email, phone, institution,
            panel, diet, special
        ])
        wb.save(EXCEL_FILE)

        # Email to participant
        send_email(
            email,
            "Accounting Conclave 2025 - Registration Confirmation",
            f"""
            <p>Hello {name},</p>
            <p>Thank you for registering for Accounting Conclave 2025.</p>
            <p><b>Details:</b></p>
            <ul>
                <li>Student Type: {student_type}</li>
                <li>Ticket Type: {ticket_type}</li>
                <li>Price: ₹{price}</li>
                <li>Panel Preference: {panel}</li>
                <li>Dietary Preference: {diet}</li>
            </ul>
            <p><b>Date:</b> 30th August 2025<br>
            <b>Venue:</b> Ahmedabad University</p>
            <p>We look forward to seeing you!</p>
            <p>Best Regards,<br>Organizing Team</p>
            """
        )

        # Email to admin
        send_email(
            EMAIL_RECEIVER,
            "New Registration - Accounting Conclave 2025",
            f"""
            <p><b>New Registration:</b></p>
            <ul>
                <li>Student Type: {student_type}</li>
                <li>Ticket Type: {ticket_type}</li>
                <li>Price: ₹{price}</li>
                <li>Name: {name}</li>
                <li>Email: {email}</li>
                <li>Phone: {phone}</li>
                <li>Institution: {institution}</li>
                <li>Panel Preference: {panel}</li>
                <li>Dietary Preference: {diet}</li>
                <li>Special Requirements: {special}</li>
            </ul>
            """
        )

        return redirect(url_for("success"))

    return render_template("register.html")

@app.route("/success")
def success():
    return render_template("success.html")

if __name__ == "__main__":
    app.run(debug=True)
